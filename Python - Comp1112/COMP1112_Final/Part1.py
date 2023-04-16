# COMP1112 Final Exam
# Part1.py
# Name: Dmitry Knyazhevskiy
# Student #: 1203665

'''
Takes a text file with a list of websites
Then sends them an HTTP request
Then writes the name and code that the request returns to a spreadsheet
'''
"""
Changes:
1. I would automatically have it append the xlsx file format to the name you provide, so you dont have write out the entire thing
2. Some minor parts can be consolidated, like having the 'urlLen = len(urls) be inside the for loop
"""

import requests
from openpyxl import Workbook

#Asks the user to input names into the console, one for the starting file and one for the final file
fnameIn = input("Please provide an input file name: ")
fnameOut = input("Please provide an output file name: ")

#Opens the starter file in read mode
fIn = open(fnameIn, "r")
contents = fIn.read()
#Then splits each of the entries by the newline character and saves each entry as a new item in a list called urls
urls = contents.split("\n")
#Creates an empty dictionary for the results of the HTTP requests
results = {}
#Gets the number of items in the urls list
urlLen = len(urls)

#For each item in the urls list, it sends out an http request then saves the code it recieves back in the results dictionary with the key being the website
for i in range(urlLen):
    url = urls[i]
    result = requests.get(url)
    results[url] = result.status_code

#Creates a new workbook called wb and sets it as the active workbook, and set row one as the current row
wb = Workbook()
ws = wb.active
row = 1

#For each item in the results dictionary it does
#Prints out the website
#In the first column of the selected row it writes the website, then in the second column it writes the reulting code
#Then it adds one to the row variable
for result in results:
    print(result)
    ws['A' + str(row)] = result
    ws['B' + str(row)] = results[result]
    row += 1

#Saves the file with the previously inputted output file name
wb.save(filename=fnameOut)