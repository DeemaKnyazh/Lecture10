# COMP1112 Final Exam
# Part4.py
# Name: Dmitry Knyazhevskiy
# Student #: 1203665

from openpyxl import Workbook

def makeWorkbook(name, testList):
    wb = Workbook() #Create a new workbook called wb
    sheet1 = wb.active #Sets the first sheet as the active one
    sheet1.title = "Sheet" #Titles the first sheet as Sheet 

    alpha = ["A","B","C","D","E","F","G","H","I","J","K","L","M"] #List of letters for labelling columns
    col1Len = []#Empty list for later use
    rowLen = len(testList)#Number of lists in the list of lists

    #For loop that finds the longest list
    for i in testList:
        col1Len.append(len(i))
        maxLen = max(col1Len)

    #For loop that uses the length of the longest list to populate the letters at the top of the spreadsheet
    for i in range(maxLen):
        sheet1.cell(row=1,column=i+1).value = alpha[i]
    
    #Nested for loop that goes through each column of each row individually and saves to that cell the corresponding value in the list
    for i in range(rowLen):
        col2Len = len(testList[i])
        for k in range(col2Len):
            sheet1.cell(row=i+1,column=k+1).value = testList[i][k]

    wb.save(name + '.xlsx') # Saves the edited workbook to the spreadsheet

#List for testing purposes and function called with the name test and test list
myList = [["1","2","3","4"],[4,"Test",6,"Eight"],[7.3,"35", 4.2]]
makeWorkbook('test', myList)